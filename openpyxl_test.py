from openpyxl import load_workbook

wb = load_workbook(filename='C:\\Users\\User\\Desktop\\test.xlsx')
res = len(wb.sheetnames)

data_keywords = []
data_variables = []
file = open("C:\\Users\\User\\Desktop\\abhi.txt", "w")


def keyword(page_cd1, page_desc1):
    return "Search-PageCode-FIN-" + page_cd1 + "\n\t" + "Search-PageCode-FIN\t" + "${" + page_cd1 + "}\t" + "${" + page_desc1 + "}\n\n"


def variables(page_cd1, page_desc1):
    return "${" + page_cd1 + "}\t" + page_cd1 + "\n${" + page_desc1 + "}\t" + page_desc1 + "\n\n"


for i in range(res):
    sheetNum = 'Sheet' + str(i+1)
    sheet_ranges = wb[sheetNum]
    rows = sheet_ranges.max_row
    for j in range(rows-1):
        cellA = 'A' + str(j+2)
        cellB = 'B' + str(j+2)
        page_cd = sheet_ranges[cellA].value
        page_desc = sheet_ranges[cellB].value

        data_variables.append(variables(page_cd,page_desc))
        data_keywords.append(keyword(page_cd, page_desc))

file.write("***Variables***\n")
file.writelines(data_variables)
file.write("***Keywords***\n")
file.writelines(data_keywords)

file.close()
